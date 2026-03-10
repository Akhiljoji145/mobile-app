import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os

def create_test_cases():
    # Define the directory and file path
    directory = r'e:\test_case_of_e_bus'
    file_path = os.path.join(directory, 'login_test_cases.xlsx')

    # Create directory if it doesn't exist
    if not os.path.exists(directory):
        os.makedirs(directory)

    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Login Test Cases"

    # Define headers
    headers = ["Test Case ID", "Test Scenario", "Test Steps", "Input Data", "Expected Result", "Status"]
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Define test cases
    test_cases = [
        ["TC_001", "Successful Login", 
         "1. Enter valid username\n2. Enter valid password\n3. Click 'Sign In'", 
         "Username: valid_user\nPassword: valid_password", 
         "User is redirected to the dashboard/home screen.", ""],
        
        ["TC_002", "Login with Empty Fields", 
         "1. Leave username and password empty\n2. Click 'Sign In'", 
         "Username: [Empty]\nPassword: [Empty]", 
         "Alert 'Missing Info' is displayed: 'Please enter both username and password.'", ""],
        
        ["TC_003", "Login with Missing Password", 
         "1. Enter a valid username\n2. Leave password empty\n3. Click 'Sign In'", 
         "Username: some_user\nPassword: [Empty]", 
         "Alert 'Missing Info' is displayed.", ""],
        
        ["TC_004", "Login with Invalid Credentials", 
         "1. Enter incorrect username/password\n2. Click 'Sign In'", 
         "Username: wrong_user\nPassword: wrong_password", 
         "Alert 'Login Failed' is displayed with an error message.", ""],
        
        ["TC_005", "Blocked Account Scenario", 
         "1. Enter credentials for a blocked account\n2. Click 'Sign In'", 
         "Username: blocked_user\nPassword: some_password", 
         "A red toast notification appears at the top indicating the account is blocked.", ""],
        
        ["TC_006", "Forgot Password Navigation", 
         "1. Click on 'Forgot Password?' link", 
         "N/A", 
         "User is navigated to the 'ForgotPassword' screen.", ""],
        
        ["TC_007", "Loading Indicator Visibility", 
         "1. Enter valid credentials\n2. Click 'Sign In'", 
         "N/A", 
         "Login button text is replaced by an ActivityIndicator (spinner) while the request is in progress.", ""],
        
        ["TC_008", "UI Elements Verification", 
         "1. Observe the login screen", 
         "N/A", 
         "Logo, title 'EduTransit', and subtitle 'Welcome Back...' are correctly displayed.", ""],
    ]

    # Add test cases to the sheet
    for tc in test_cases:
        ws.append(tc)

    # Adjust column widths
    column_widths = [15, 30, 50, 40, 50, 10]
    for i, column_width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width

    # Apply text wrapping
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Save the workbook
    wb.save(file_path)
    print(f"Successfully created: {file_path}")

if __name__ == "__main__":
    create_test_cases()
